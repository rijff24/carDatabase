"""
Helpers for importing data from CSV and Excel files.

These functions handle the parsing and validation of imported data.
"""

import pandas as pd
import logging
from app import db
from werkzeug.datastructures import FileStorage
from typing import Dict, List, Any, Union
from datetime import datetime
from app.models.car import Car, VehicleMake, VehicleModel, VehicleColor
from app.models.dealer import Dealer
from app.models.stand import Stand
from sqlalchemy.exc import SQLAlchemyError
import re
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from io import BytesIO


def import_cars(file: FileStorage) -> Dict[str, Any]:
    """
    Import cars from a CSV or Excel file.
    
    Args:
        file: Uploaded file object
        
    Returns:
        dict: Results of the import with success and failure counts
    """
    logging.info("Car import started")
    success_count = 0
    fail_count = 0
    errors = []
    warnings = []
    
    try:
        # Validate file format and read into DataFrame
        if not file or not file.filename:
            errors.append("No file provided")
            return {"success": 0, "failed": 0, "errors": errors, "warnings": warnings}
        
        # Check file extension
        file_ext = os.path.splitext(file.filename)[1][1:].lower()
        if file_ext not in ['csv', 'xlsx']:
            errors.append(f"Unsupported file format: {file_ext}. Please use CSV or Excel (.xlsx)")
            return {"success": 0, "failed": 0, "errors": errors, "warnings": warnings}
        
        # Read file into DataFrame
        try:
            if file_ext == 'csv':
                df = pd.read_csv(file.stream)
            else:  # xlsx
                df = pd.read_excel(file.stream)
        except Exception as e:
            errors.append(f"Error reading file: {str(e)}")
            return {"success": 0, "failed": 0, "errors": errors, "warnings": warnings}
        
        # Check if DataFrame is empty
        if df.empty:
            errors.append("File contains no data")
            return {"success": 0, "failed": 0, "errors": errors, "warnings": warnings}
        
        # Get existing data for relationships to avoid duplicates
        makes_dict = {m.name.lower(): m for m in VehicleMake.query.all()}
        models_dict = {(m.name.lower(), m.make_id): m for m in VehicleModel.query.all()}
        colors_dict = {c.name.lower(): c for c in VehicleColor.query.all()}
        stands_dict = {s.stand_name.lower(): s for s in Stand.query.all()}
        
        # Get a default stand for missing stand relationships
        default_stand = Stand.query.first()
        default_stand_id = default_stand.stand_id if default_stand else None
        
        # Pre-process input data to handle duplicates in relationships
        # Especially for cases where multiple rows reference the same stand by name
        stand_name_mapping = {}
        
        # First scan to identify and cache relationships
        for index, row in df.iterrows():
            # Create mappings for stand names
            if 'Stand' in df.columns and not pd.isna(row.get('Stand')) and str(row.get('Stand')).strip():
                stand_name = str(row.get('Stand')).strip().lower()
                if stand_name not in stand_name_mapping:
                    # Look for existing stand
                    stand = next((s for s_name, s in stands_dict.items() if s_name == stand_name), None)
                    if stand:
                        stand_name_mapping[stand_name] = stand.stand_id
                    elif default_stand_id:
                        stand_name_mapping[stand_name] = default_stand_id
                        warnings.append(f"Stand '{row.get('Stand')}' not found, using default stand")
        
        # Process each row
        for index, row in df.iterrows():
            row_data = {}
            row_warnings = []
            
            # Extract values with validation
            try:
                # Licence Number (required)
                if 'Licence Number' in df.columns and not pd.isna(row.get('Licence Number')) and str(row.get('Licence Number')).strip():
                    licence_number = str(row.get('Licence Number')).strip().upper()
                    
                    # Check for duplicate licence number
                    existing_car = Car.query.filter_by(licence_number=licence_number).first()
                    if existing_car:
                        fail_count += 1
                        errors.append(f"Row {index+2}: Car with licence number '{licence_number}' already exists. Skipping.")
                        continue
                    
                    row_data['licence_number'] = licence_number
                else:
                    fail_count += 1
                    errors.append(f"Row {index+2}: Licence Number is required. Skipping.")
                    continue
                
                # Registration Number (required)
                if 'Registration Number' in df.columns and not pd.isna(row.get('Registration Number')) and str(row.get('Registration Number')).strip():
                    row_data['registration_number'] = str(row.get('Registration Number')).strip().upper()
                else:
                    row_data['registration_number'] = licence_number  # Use licence number if not provided
                    row_warnings.append(f"No Registration Number provided, using Licence Number")
                
                # Make (required for vehicle name calculation)
                if 'Make' in df.columns and not pd.isna(row.get('Make')) and str(row.get('Make')).strip():
                    make_str = str(row.get('Make')).strip()
                    
                    # Handle Make relationship - find or create
                    sanitized_make = VehicleMake.sanitize_name(make_str)
                    make_obj = None
                    
                    # Look for existing make
                    if sanitized_make.lower() in makes_dict:
                        make_obj = makes_dict[sanitized_make.lower()]
                    else:
                        make_obj = VehicleMake(name=sanitized_make)
                        db.session.add(make_obj)
                        db.session.flush()  # Get ID without committing
                        makes_dict[sanitized_make.lower()] = make_obj
                    
                    row_data['vehicle_make'] = make_obj.name
                else:
                    fail_count += 1
                    errors.append(f"Row {index+2}: Make is required. Skipping.")
                    continue
                
                # Model (required for vehicle name calculation)
                if 'Model' in df.columns and not pd.isna(row.get('Model')) and str(row.get('Model')).strip():
                    model_str = str(row.get('Model')).strip()
                    
                    # Handle Model relationship - find or create
                    sanitized_model = VehicleModel.sanitize_name(model_str)
                    model_obj = None
                    
                    # Look for existing model
                    if (sanitized_model.lower(), make_obj.id) in models_dict:
                        model_obj = models_dict[(sanitized_model.lower(), make_obj.id)]
                    else:
                        model_obj = VehicleModel(name=sanitized_model, make_id=make_obj.id)
                        db.session.add(model_obj)
                        db.session.flush()  # Get ID without committing
                        models_dict[(sanitized_model.lower(), make_obj.id)] = model_obj
                    
                    row_data['vehicle_model'] = model_obj.name
                else:
                    fail_count += 1
                    errors.append(f"Row {index+2}: Model is required. Skipping.")
                    continue
                
                # Year (required for vehicle name calculation)
                if 'Year' in df.columns and not pd.isna(row.get('Year')):
                    try:
                        year = int(row.get('Year'))
                        # Basic year validation
                        current_year = datetime.now().year
                        if year < 1900 or year > current_year + 1:
                            fail_count += 1
                            errors.append(f"Row {index+2}: Year {year} is invalid (must be between 1900 and {current_year+1}). Skipping.")
                            continue
                        row_data['year'] = year
                    except (ValueError, TypeError):
                        fail_count += 1
                        errors.append(f"Row {index+2}: Year must be a valid number. Skipping.")
                        continue
                else:
                    fail_count += 1
                    errors.append(f"Row {index+2}: Year is required. Skipping.")
                    continue
                
                # Color (required)
                if 'Colour' in df.columns and not pd.isna(row.get('Colour')) and str(row.get('Colour')).strip():
                    color = str(row.get('Colour')).strip()
                    
                    # Handle Color relationship - find or create
                    color_sanitized = VehicleColor.sanitize_name(color)
                    color_obj = None
                    
                    # Look for existing color
                    if color_sanitized.lower() in colors_dict:
                        color_obj = colors_dict[color_sanitized.lower()]
                    else:
                        color_obj = VehicleColor(name=color_sanitized)
                        db.session.add(color_obj)
                        db.session.flush()  # Get ID without committing
                        colors_dict[color_sanitized.lower()] = color_obj
                    
                    row_data['colour'] = color_sanitized
                else:
                    fail_count += 1
                    errors.append(f"Row {index+2}: Colour is required. Skipping.")
                    continue
                
                # Dekra condition (required)
                if 'Dekra Condition' in df.columns and not pd.isna(row.get('Dekra Condition')) and str(row.get('Dekra Condition')).strip():
                    row_data['dekra_condition'] = str(row.get('Dekra Condition')).strip()
                else:
                    row_data['dekra_condition'] = "Unknown"  # Changed from "Good" to "Unknown"
                    row_warnings.append("No dekra condition provided, using 'Unknown'")
                
                # Purchase price (required)
                if 'Purchase Price' in df.columns and not pd.isna(row.get('Purchase Price')):
                    try:
                        purchase_price = float(row.get('Purchase Price'))
                        if purchase_price < 0:
                            fail_count += 1
                            errors.append(f"Row {index+2}: Purchase price cannot be negative. Skipping.")
                            continue
                        row_data['purchase_price'] = purchase_price
                    except (ValueError, TypeError):
                        fail_count += 1
                        errors.append(f"Row {index+2}: Purchase price must be a valid number. Skipping.")
                        continue
                else:
                    fail_count += 1
                    errors.append(f"Row {index+2}: Purchase Price is required. Skipping.")
                    continue
                
                # Recon cost (optional)
                if 'Recon Cost' in df.columns and not pd.isna(row.get('Recon Cost')):
                    try:
                        recon_cost = float(row.get('Recon Cost'))
                        if recon_cost < 0:
                            row_warnings.append("Negative recon cost, using 0.0")
                            recon_cost = 0.0
                        row_data['recon_cost'] = recon_cost
                    except (ValueError, TypeError):
                        row_warnings.append("Invalid recon cost format, using 0.0")
                        row_data['recon_cost'] = 0.0
                
                # Final cost price (optional)
                if 'Final Cost Price' in df.columns and not pd.isna(row.get('Final Cost Price')):
                    try:
                        final_cost_price = float(row.get('Final Cost Price'))
                        if final_cost_price < 0:
                            recon_cost = row_data.get('recon_cost', 0.0)
                            final_cost_price = purchase_price + recon_cost
                            row_warnings.append(f"Negative final cost price, calculating as purchase + recon = {final_cost_price}")
                        row_data['final_cost_price'] = final_cost_price
                    except (ValueError, TypeError):
                        recon_cost = row_data.get('recon_cost', 0.0)
                        row_data['final_cost_price'] = purchase_price + recon_cost
                        row_warnings.append(f"Invalid final cost price format, calculating as purchase + recon = {row_data['final_cost_price']}")
                else:
                    # If not provided, calculate from purchase price and recon cost
                    recon_cost = row_data.get('recon_cost', 0.0)
                    row_data['final_cost_price'] = purchase_price + recon_cost
                    row_warnings.append(f"No final cost price provided, calculating as purchase + recon = {row_data['final_cost_price']}")
                
                # Sell Price (optional)
                if 'Sell Price' in df.columns and not pd.isna(row.get('Sell Price')):
                    try:
                        sell_price = float(row.get('Sell Price'))
                        if sell_price >= 0:
                            row_data['sale_price'] = sell_price
                    except (ValueError, TypeError):
                        row_warnings.append("Invalid sell price format, ignoring")
                
                # Date bought (required)
                if 'Date Bought' in df.columns and not pd.isna(row.get('Date Bought')):
                    try:
                        if isinstance(row.get('Date Bought'), str):
                            row_data['date_bought'] = datetime.strptime(row.get('Date Bought'), '%Y-%m-%d').date()
                        elif isinstance(row.get('Date Bought'), datetime):
                            row_data['date_bought'] = row.get('Date Bought').date()
                        else:
                            row_data['date_bought'] = datetime.now().date()
                            row_warnings.append("Invalid date bought format, using today's date")
                    except (ValueError, TypeError):
                        row_data['date_bought'] = datetime.now().date()
                        row_warnings.append("Invalid date bought format, using today's date")
                else:
                    row_data['date_bought'] = datetime.now().date()
                    row_warnings.append("No date bought provided, using today's date")
                
                # Date added to stand (optional)
                if 'Date Added To Stand' in df.columns and not pd.isna(row.get('Date Added To Stand')):
                    try:
                        if isinstance(row.get('Date Added To Stand'), str):
                            row_data['date_added_to_stand'] = datetime.strptime(row.get('Date Added To Stand'), '%Y-%m-%d').date()
                        elif isinstance(row.get('Date Added To Stand'), datetime):
                            row_data['date_added_to_stand'] = row.get('Date Added To Stand').date()
                    except (ValueError, TypeError):
                        row_warnings.append("Invalid date added to stand format, leaving empty")
                elif 'date_sold' in row_data:
                    # If car has been sold but no date added to stand, set to current date
                    row_data['date_added_to_stand'] = datetime.now().date()
                    row_warnings.append("Car is sold but no date added to stand provided, using today's date")
                
                # ===== Adding Sale Information =====
                # Date sold (optional)
                date_sold = None
                if 'Date Sold' in df.columns and not pd.isna(row.get('Date Sold')):
                    try:
                        if isinstance(row.get('Date Sold'), str):
                            date_sold = datetime.strptime(row.get('Date Sold'), '%Y-%m-%d').date()
                        elif isinstance(row.get('Date Sold'), datetime):
                            date_sold = row.get('Date Sold').date()
                        row_data['date_sold'] = date_sold
                    except (ValueError, TypeError):
                        row_warnings.append("Invalid date sold format, treating as not sold")
                
                # Sale price (optional)
                sale_price = None
                if 'Sale Price' in df.columns and not pd.isna(row.get('Sale Price')):
                    try:
                        sale_price = float(row.get('Sale Price'))
                        if sale_price < 0:
                            row_warnings.append("Negative sale price, treating as not sold")
                        else:
                            row_data['sale_price'] = sale_price
                    except (ValueError, TypeError):
                        row_warnings.append("Invalid sale price format, treating as not sold")
                
                # Handle sold status logic:
                # If either date_sold or sale_price is provided, the car is considered sold
                if (date_sold is not None) or (sale_price is not None):
                    row_data['repair_status'] = 'Sold'
                    
                    # If one is missing but the other is present, keep the provided one as is
                    if date_sold is None and sale_price is not None:
                        row_warnings.append("Sale price provided without date sold, marking as sold with no date")
                    
                    if sale_price is None and date_sold is not None:
                        row_warnings.append("Date sold provided without sale price, marking as sold with no price")
                else:
                    # If neither is provided, car is not sold
                    # We'll set the status below, but make sure these fields are not set
                    if 'date_sold' in row_data:
                        del row_data['date_sold']
                    if 'sale_price' in row_data:
                        del row_data['sale_price']
                
                # Refuel cost (optional)
                if 'Refuel Cost' in df.columns and not pd.isna(row.get('Refuel Cost')):
                    try:
                        refuel_cost = float(row.get('Refuel Cost'))
                        if refuel_cost < 0:
                            row_warnings.append("Negative refuel cost, using 0.0")
                            refuel_cost = 0.0
                        row_data['refuel_cost'] = refuel_cost
                    except (ValueError, TypeError):
                        row_warnings.append("Invalid refuel cost format, using 0.0")
                        row_data['refuel_cost'] = 0.0
                else:
                    row_data['refuel_cost'] = 0.0
                
                # Current location (optional with default)
                if 'Current Location' in df.columns and not pd.isna(row.get('Current Location')) and str(row.get('Current Location')).strip():
                    row_data['current_location'] = str(row.get('Current Location')).strip()
                else:
                    row_data['current_location'] = "Showroom"  # Default
                    row_warnings.append("No current location provided, using 'Showroom'")
                
                # Status (optional with default, or from sale status)
                if 'repair_status' not in row_data:  # If not already set by sale condition
                    if 'Status' in df.columns and not pd.isna(row.get('Status')) and str(row.get('Status')).strip():
                        status = str(row.get('Status')).strip()
                        valid_statuses = ['Available', 'In Repairs', 'Sold', 'Reserved']
                        if status in valid_statuses:
                            row_data['repair_status'] = status
                        else:
                            row_data['repair_status'] = 'Available'
                            row_warnings.append(f"Invalid status '{status}', using 'Available'")
                    elif 'recon_cost' in row_data and row_data['recon_cost'] > 0:
                        # If car has recon cost but no status, set to "Waiting for Repairs"
                        row_data['repair_status'] = 'Waiting for Repairs'
                        row_warnings.append("Car has recon cost but no status provided, using 'Waiting for Repairs'")
                    else:
                        row_data['repair_status'] = 'Available'  # Default
                        row_warnings.append("No status provided, using 'Available'")
                
                # Source (optional with default)
                if 'Source' in df.columns and not pd.isna(row.get('Source')) and str(row.get('Source')).strip():
                    row_data['source'] = str(row.get('Source')).strip()
                else:
                    row_data['source'] = "Import"  # Default
                    row_warnings.append("No source provided, using 'Import'")
                
                # Stand relationship (optional with default)
                if 'Stand' in df.columns and not pd.isna(row.get('Stand')) and str(row.get('Stand')).strip():
                    stand_name = str(row.get('Stand')).strip().lower()
                    if stand_name in stand_name_mapping:
                        row_data['stand_id'] = stand_name_mapping[stand_name]
                    elif default_stand_id:
                        row_data['stand_id'] = default_stand_id
                        row_warnings.append(f"Stand '{row.get('Stand')}' not found, using default stand")
                elif default_stand_id:
                    row_data['stand_id'] = default_stand_id
                    row_warnings.append("No stand provided, using 'Unknown'")
                
                # Vehicle name (derived or provided)
                if 'Vehicle Name' in df.columns and not pd.isna(row.get('Vehicle Name')) and str(row.get('Vehicle Name')).strip():
                    row_data['vehicle_name'] = str(row.get('Vehicle Name')).strip()
                else:
                    row_data['vehicle_name'] = f"{year} {make_obj.name} {model_obj.name}"
                    row_warnings.append(f"No vehicle name provided, using '{row_data['vehicle_name']}'")
                
                # Create the new car
                new_car = Car(**row_data)
                db.session.add(new_car)
                db.session.commit()
                success_count += 1
                
                # Log any warnings
                if row_warnings:
                    warnings.append(f"Row {index+2} imported with warnings: {'; '.join(row_warnings)}")
                
            except Exception as e:
                db.session.rollback()
                fail_count += 1
                errors.append(f"Row {index+2}: Error processing data: {str(e)}")
                logging.error(f"Error processing row {index+2}: {str(e)}")
    except Exception as e:
        logging.error(f"Overall import error: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        errors.append(f"Import process error: {str(e)}")
    
    return {
        "success": success_count,
        "failed": fail_count,
        "errors": errors,
        "warnings": warnings
    }


def import_repairs(file: FileStorage) -> Dict[str, Any]:
    """
    Import repairs from a CSV or Excel file.
    
    Args:
        file: Uploaded file object
        
    Returns:
        dict: Results of the import with success and failure counts
    """
    logging.info("Repairs import started")
    # The implementation will be done later
    # For now, return a stub response
    return {
        "success": 0,
        "failed": 0,
        "errors": []
    }


def import_sales(file: FileStorage) -> Dict[str, Any]:
    """
    Import sales from a CSV or Excel file.
    
    Args:
        file: Uploaded file object
        
    Returns:
        dict: Results of the import with success and failure counts
    """
    logging.info("Sales import started")
    # The implementation will be done later
    # For now, return a stub response
    return {
        "success": 0,
        "failed": 0,
        "errors": []
    }


def import_dealers(file: FileStorage) -> Dict[str, Any]:
    """
    Import dealers from a CSV or Excel file.
    
    Args:
        file: Uploaded file object
        
    Returns:
        dict: Results of the import with success and failure counts
    """
    logging.info("Dealers import started")
    # The implementation will be done later
    # For now, return a stub response
    return {
        "success": 0,
        "failed": 0,
        "errors": []
    }


def import_parts(file: FileStorage) -> Dict[str, Any]:
    """
    Import parts from a CSV or Excel file.
    
    Args:
        file: Uploaded file object
        
    Returns:
        dict: Results of the import with success and failure counts
    """
    logging.info("Parts import started")
    # The implementation will be done later
    # For now, return a stub response
    return {
        "success": 0,
        "failed": 0,
        "errors": []
    }


def import_stands(file: FileStorage) -> Dict[str, Any]:
    """
    Import stands from a CSV or Excel file.
    
    Args:
        file: Uploaded file object
        
    Returns:
        dict: Results of the import with success and failure counts
    """
    logging.info("Stands import started")
    # The implementation will be done later
    # For now, return a stub response
    return {
        "success": 0,
        "failed": 0,
        "errors": []
    }


def create_sample_template(entity_type: str) -> BytesIO:
    """
    Generate and return a sample template Excel file for importing data.
    
    Args:
        entity_type: Type of entity (cars, repairs, sales, etc.)
        
    Returns:
        BytesIO: Excel file buffer
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{entity_type.capitalize()} Import Template"
    
    fill_header = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    fill_required = PatternFill(start_color="F8C471", end_color="F8C471", fill_type="solid")
    fill_optional = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    
    # Create header style
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define headers and column specs based on entity type
    headers = []
    col_widths = []
    
    if entity_type == 'cars':
        headers = [
            {"name": "Licence Number", "comment": "Required. Licence number of the vehicle. Must be unique for each car.", "type": "required", "width": 20},
            {"name": "Registration Number", "comment": "Optional. Registration number of the vehicle. Default: Same as licence number", "type": "optional", "width": 20},
            {"name": "Vehicle Name", "comment": "Optional. Full name of the vehicle. Default: Year Make Model", "type": "optional", "width": 25},
            {"name": "Make", "comment": "Required. Vehicle manufacturer (e.g., Toyota, Ford).", "type": "required", "width": 15},
            {"name": "Model", "comment": "Required. Vehicle model (e.g., Corolla, F-150).", "type": "required", "width": 15},
            {"name": "Year", "comment": "Required. Year of manufacture (1900-current).", "type": "required", "width": 8},
            {"name": "Colour", "comment": "Required. Vehicle color.", "type": "required", "width": 12},
            {"name": "Dekra Condition", "comment": "Optional. Condition of the vehicle. Default: 'Unknown'", "type": "optional", "width": 15},
            {"name": "Purchase Price", "comment": "Required. Amount paid to acquire the vehicle.", "type": "required", "width": 15},
            {"name": "Recon Cost", "comment": "Optional. Reconditioning cost for the vehicle. Default: 0", "type": "optional", "width": 15},
            {"name": "Final Cost Price", "comment": "Optional. Final cost including purchase and recon. Default: Purchase Price + Recon Cost", "type": "optional", "width": 15},
            {"name": "Sell Price", "comment": "Optional. Suggested selling price for the vehicle.", "type": "optional", "width": 15},
            {"name": "Date Bought", "comment": "Optional. Date the vehicle was purchased (YYYY-MM-DD). Default: today's date", "type": "optional", "width": 15},
            {"name": "Date Added To Stand", "comment": "Optional. Date the vehicle was added to a stand (YYYY-MM-DD). If car is sold but this is empty, defaults to current date", "type": "optional", "width": 18},
            {"name": "Date Sold", "comment": "Optional. Date the vehicle was sold (YYYY-MM-DD). If either Date Sold or Sale Price is provided, the car is marked as sold.", "type": "optional", "width": 15},
            {"name": "Sale Price", "comment": "Optional. Amount the vehicle was sold for. If either Date Sold or Sale Price is provided, the car is marked as sold.", "type": "optional", "width": 15},
            {"name": "Refuel Cost", "comment": "Optional. Cost to refuel the vehicle. Default: 0", "type": "optional", "width": 12},
            {"name": "Current Location", "comment": "Optional. Current location of the vehicle. Default: 'Showroom'", "type": "optional", "width": 15},
            {"name": "Status", "comment": "Optional. Vehicle status (Available, In Repairs, Reserved). Default: 'Available'. Automatically set to 'Sold' if car is sold, or 'Waiting for Repairs' if recon cost is provided but no status.", "type": "optional", "width": 12},
            {"name": "Source", "comment": "Optional. Source of the vehicle. Default: 'Import'", "type": "optional", "width": 15},
            {"name": "Stand", "comment": "Optional. Stand name. If empty, defaults to 'Unknown'.", "type": "optional", "width": 20},
        ]
    elif entity_type == 'repairs':
        headers = [
            {"name": "Car Licence Number", "comment": "Required. Licence Number of the car to repair.", "type": "required", "width": 20},
            {"name": "Description", "comment": "Required. Description of repairs needed or performed.", "type": "required", "width": 40},
            {"name": "Repair Cost", "comment": "Required. Cost of repair work (numeric value).", "type": "required", "width": 15},
            {"name": "Provider", "comment": "Optional. Service provider who performed the repairs.", "type": "optional", "width": 20},
            {"name": "Start Date", "comment": "Optional. Date repairs started (YYYY-MM-DD).", "type": "optional", "width": 15},
            {"name": "End Date", "comment": "Optional. Date repairs completed (YYYY-MM-DD).", "type": "optional", "width": 15},
            {"name": "Notes", "comment": "Optional. Additional notes about the repair.", "type": "optional", "width": 40},
        ]
        
    elif entity_type == 'sales':
        headers = [
            {"name": "Car Licence Number", "comment": "Required. Licence Number of the car sold.", "type": "required", "width": 20},
            {"name": "Sale Date", "comment": "Required. Date of sale (YYYY-MM-DD).", "type": "required", "width": 15},
            {"name": "Sale Price", "comment": "Required. Sale price (numeric value).", "type": "required", "width": 15},
            {"name": "Customer Name", "comment": "Required. Name of customer who purchased the car.", "type": "required", "width": 25},
            {"name": "Customer Contact", "comment": "Optional. Contact information for the customer.", "type": "optional", "width": 25},
            {"name": "Payment Method", "comment": "Optional. Method of payment (Cash, Credit, Finance, etc.).", "type": "optional", "width": 15},
            {"name": "Notes", "comment": "Optional. Additional notes about the sale.", "type": "optional", "width": 40},
        ]
        
    elif entity_type == 'dealers':
        headers = [
            {"name": "Dealer Name", "comment": "Required. Name of the dealer.", "type": "required", "width": 25},
            {"name": "Contact Info", "comment": "Required. Contact information for the dealer.", "type": "required", "width": 25},
            {"name": "Address", "comment": "Optional. Physical address of the dealer.", "type": "optional", "width": 40},
            {"name": "Status", "comment": "Optional. Status of the dealer (Active, Inactive, Suspended). Default: Active", "type": "optional", "width": 15},
        ]
        
    elif entity_type == 'parts':
        headers = [
            {"name": "Part Name", "comment": "Required. Name of the part.", "type": "required", "width": 25},
            {"name": "Description", "comment": "Optional. Description of the part.", "type": "optional", "width": 40},
            {"name": "Manufacturer", "comment": "Optional. Manufacturer of the part.", "type": "optional", "width": 20},
            {"name": "Price", "comment": "Required. Price of the part (numeric value).", "type": "required", "width": 15},
            {"name": "SKU", "comment": "Optional. Stock Keeping Unit or part number.", "type": "optional", "width": 15},
            {"name": "Quantity", "comment": "Optional. Quantity in stock. Default: 0", "type": "optional", "width": 10},
        ]
        
    elif entity_type == 'stands':
        headers = [
            {"name": "Stand Name", "comment": "Required. Name of the stand.", "type": "required", "width": 25},
            {"name": "Location", "comment": "Required. Location of the stand.", "type": "required", "width": 30},
            {"name": "Capacity", "comment": "Optional. Vehicle capacity of the stand. Default: 10", "type": "optional", "width": 10},
            {"name": "Additional Info", "comment": "Optional. Additional information about the stand.", "type": "optional", "width": 40},
        ]
    
    # Add headers and format them
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header["name"]
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Add column width
        ws.column_dimensions[get_column_letter(col_idx)].width = header["width"]
        
        # Add comments with instructions
        comment = Comment(header["comment"], "Template Generator")
        comment.width = 400
        comment.height = 100
        cell.comment = comment
        
        # Add background color based on required/optional
        if header["type"] == "required":
            ws.cell(row=2, column=col_idx).fill = fill_required
        else:
            ws.cell(row=2, column=col_idx).fill = fill_optional
    
    # Add some sample data based on entity type
    if entity_type == 'cars':
        # First sample vehicle (complete)
        ws.append([
            "ABC123456789",       # Licence Number (required)
            "XYZ987654321",       # Registration Number
            "2020 Toyota Corolla Premium", # Vehicle Name
            "Toyota",             # Make (required)
            "Corolla",            # Model (required)
            2020,                 # Year (required)
            "Blue",               # Colour (required)
            "Excellent",          # Dekra Condition
            15000,                # Purchase Price (required)
            500,                  # Recon Cost
            15500,                # Final Cost Price
            18000,                # Sell Price
            datetime.now().date(), # Date Bought
            datetime.now().date(), # Date Added To Stand
            "",                   # Date Sold (empty = not sold)
            "",                   # Sale Price (empty = not sold)
            50,                   # Refuel Cost
            "Showroom",           # Current Location
            "Available",          # Status
            "Auction",            # Source
            "Main Stand",         # Stand
        ])
        
        # Second sample vehicle with minimal required info and sold status
        ws.append([
            "DEF456789012",       # Licence Number (required)
            "",                   # Registration Number (will use licence number)
            "",                   # Vehicle Name (will be generated)
            "Ford",               # Make (required)
            "F-150",              # Model (required)
            2022,                 # Year (required)
            "Red",                # Colour (required)
            "",                   # Dekra Condition
            12500,                # Purchase Price (required)
            "",                   # Recon Cost
            "",                   # Final Cost Price
            13500,                # Sell Price
            "",                   # Date Bought
            "",                   # Date Added To Stand
            datetime.now().date(), # Date Sold (this will mark the car as sold)
            15000,                # Sale Price
            "",                   # Refuel Cost
            "",                   # Current Location
            "",                   # Status (will be auto-set to Sold)
            "",                   # Source
            "",                   # Stand
        ])
        
    elif entity_type == 'repairs':
        ws.append([
            "ABC123456789",       # Car Licence Number
            "Oil change and filter replacement",  # Description
            75.50,                # Repair Cost
            "Joe's Auto Shop",    # Provider
            datetime.now().date(), # Start Date
            datetime.now().date(), # End Date
            "Synthetic oil used"  # Notes
        ])
        
    elif entity_type == 'sales':
        ws.append([
            "ABC123456789",       # Car Licence Number
            datetime.now().date(), # Sale Date
            18500,                # Sale Price
            "Jane Smith",         # Customer Name
            "jane.smith@email.com, 555-1234", # Customer Contact
            "Finance",            # Payment Method
            "Customer very satisfied with purchase" # Notes
        ])
        
    elif entity_type == 'dealers':
        ws.append([
            "AutoMax Dealership", # Dealer Name
            "sales@automax.com, 555-AUTO", # Contact Info
            "123 Car Street, Autoville", # Address
            "Active"              # Status
        ])
        
    elif entity_type == 'parts':
        ws.append([
            "Oil Filter",         # Part Name
            "Standard oil filter for most Toyota vehicles", # Description
            "Toyota",             # Manufacturer
            12.99,                # Price
            "OF-TOY-001",         # SKU
            25                    # Quantity
        ])
        
    elif entity_type == 'stands':
        ws.append([
            "Main Showroom",      # Stand Name
            "1234 Display Ave, Cartown", # Location
            20,                   # Capacity
            "Premium display area with climate control" # Additional Info
        ])
    
    # Add legend at the bottom
    legend_row = len(headers) + 6  # Add some space
    ws.cell(row=legend_row, column=1).value = "Legend:"
    ws.cell(row=legend_row, column=1).font = Font(bold=True)
    
    ws.cell(row=legend_row+1, column=1).fill = fill_required
    ws.cell(row=legend_row+1, column=2).value = "Required field"
    
    ws.cell(row=legend_row+2, column=1).fill = fill_optional
    ws.cell(row=legend_row+2, column=2).value = "Optional field"
    
    if entity_type == 'cars':
        ws.cell(row=legend_row+3, column=1).value = "Note: When providing a Stand name, it must match an existing one in the system."
        ws.cell(row=legend_row+4, column=1).value = "Status can be one of: Available, In Repairs, Reserved, Waiting for Repairs"
        ws.cell(row=legend_row+5, column=1).value = "If either Date Sold or Sale Price is provided, the car will be marked as sold regardless of Status."
        ws.cell(row=legend_row+6, column=1).value = "If car has a recon cost but no status is provided, it will be marked as 'Waiting for Repairs'"
    
    # Save to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output 